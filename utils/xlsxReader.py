import pytest
import os
from openpyxl import load_workbook
from zipfile import BadZipFile

@pytest.fixture
def excel_data():
    file_path = "C:\\Users\\SharadKumar\\OneDrive - Round the clock technologies\\Desktop\\Test_Google\\test\\data.txt"
    # try:
    #     file_path = "C:\\Users\\SharadKumar\\OneDrive - Round the clock technologies\\Desktop\\Test_Google\\test\\data.xlsx"
    #     # Load the Excel workbook
    #     if os.path.exists(file_path):
    #         # Set read and write permissions
    #         os.chmod(file_path, 0o666)  # 0o666 represents read and write permissions for owner, group, and others
    #         print(f"Read and write permissions granted to '{file_path}'")
    #     else:
    #         print(f"File '{file_path}' does not exist")
    #     # permissions = os.stat("C:\\Users\\SharadKumar\\OneDrive - Round the clock technologies\\Desktop\\Test_Google\\test\\data.xlsx").st_mode
    #     workbook = load_workbook(file_path)
    #     # Assuming the data is in the first sheet
    #     sheet = workbook.active
    #     # Extract data from each row
    #     data = []
    #     for row in sheet.iter_rows(values_only=True):
    #         data.append(row)
    #     # Close the workbook
    #     workbook.close()
    #     return data
    # except BadZipFile:
    #     pytest.fail("Failed to open Excel file. The file may be corrupted or in an unsupported format.")
    try:
        with open(file_path, "r") as file:
            # Read all lines from the file
            lines = file.readlines()
            # Strip newline characters from each line and store the data
            data = [line.strip() for line in lines]
            return data
    except FileNotFoundError:
        print(f"File '{file_path}' not found")
        return None
